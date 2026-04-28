import os, re, time, requests
from openpyxl import load_workbook

API=os.getenv("PIPEDRIVE_API_TOKEN","").strip()
BASE="https://api.pipedrive.com/v1"
XLSX="/root/sdr-vps/runtime/deals_ativos.xlsx"

def only_digits(x):
    return re.sub(r"\D+","",str(x or ""))

def clean_email(x):
    x=str(x or "").strip()
    return x if "@" in x else ""

def clean_phone(x):
    d=only_digits(x)
    return d if len(d)>=8 else ""

def pd(method,path,**kw):
    params=kw.pop("params",{}) or {}
    params["api_token"]=API
    r=requests.request(method,f"{BASE}/{path}",params=params,timeout=60,**kw)
    r.raise_for_status()
    return r.json().get("data") if r.content else None

wb=load_workbook(XLSX, data_only=True)
ws=wb.active

headers=[str(c.value or "").strip() for c in ws[1]]

def find_col(name):
    for i,h in enumerate(headers):
        if name.lower() in h.lower():
            return i
    return None

deal_col=find_col("deal")
phone_col=find_col("telefone")
email_col=find_col("email")

created=0
skipped=0

for row in ws.iter_rows(min_row=2, values_only=True):
    deal_id=only_digits(row[deal_col] if deal_col is not None else "")
    if not deal_id:
        continue

    phone=clean_phone(row[phone_col]) if phone_col is not None else ""
    email=clean_email(row[email_col]) if email_col is not None else ""

    if not phone and not email:
        skipped+=1
        continue

    try:
        deal=pd("GET",f"deals/{deal_id}") or {}
        person=deal.get("person_id")
        pid=person.get("value") if isinstance(person,dict) else person
        if pid:
            continue

        org=deal.get("org_id") or {}
        org_id=org.get("value") if isinstance(org,dict) else org

        payload={"name":"Contato Comercial"}
        if org_id:
            payload["org_id"]=org_id
        if phone:
            payload["phone"]=[{"value":phone,"primary":True}]
        if email:
            payload["email"]=[{"value":email,"primary":True}]

        p=pd("POST","persons",json=payload) or {}
        new_id=p.get("id")

        if new_id:
            pd("PUT",f"deals/{deal_id}",json={"person_id":new_id})
            created+=1
            print("OK", deal_id, phone, email)

        time.sleep(0.2)

    except Exception as e:
        print("ERRO", deal_id, e)

print("FINAL created=",created," skipped=",skipped)
