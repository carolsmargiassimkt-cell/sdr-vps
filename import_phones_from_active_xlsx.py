import os, re, time, requests
from openpyxl import load_workbook

API=os.getenv("PIPEDRIVE_API_TOKEN","").strip()
BASE="https://api.pipedrive.com/v1"
XLSX="/root/sdr-vps/runtime/deals_ativos.xlsx"

if not API:
    raise SystemExit("sem PIPEDRIVE_API_TOKEN")

def digits(x):
    return re.sub(r"\D+","",str(x or ""))

def clean_phone(x):
    d=digits(x)
    return d if len(d) >= 8 else ""

def pd(method,path,**kw):
    params=kw.pop("params",{}) or {}
    params["api_token"]=API
    r=requests.request(method,f"{BASE}/{path}",params=params,timeout=60,**kw)
    r.raise_for_status()
    return r.json().get("data") if r.content else None

wb=load_workbook(XLSX, data_only=True)
ws=wb.active

created=0
skipped_existing=0
skipped_no_phone=0
errors=0

for row in ws.iter_rows(min_row=2, values_only=True):
    deal_id=digits(row[3])   # coluna 4: Negócio - ID
    org_id=digits(row[4])    # coluna 5: Organização - ID
    person_id=digits(row[10]) # coluna 11: Pessoa - ID
    person_name=str(row[11] or "").strip() # coluna 12

    phones=[]
    for idx in [15,16,17,18]:  # colunas 16-19
        p=clean_phone(row[idx])
        if p and p not in phones:
            phones.append(p)

    if not deal_id:
        continue

    if person_id:
        skipped_existing+=1
        continue

    if not phones:
        skipped_no_phone+=1
        continue

    try:
        deal=pd("GET",f"deals/{deal_id}") or {}
        current_person=deal.get("person_id")
        current_pid=current_person.get("value") if isinstance(current_person,dict) else current_person
        if current_pid:
            skipped_existing+=1
            continue

        org=deal.get("org_id") or {}
        crm_org_id=org.get("value") if isinstance(org,dict) else org
        if not org_id and crm_org_id:
            org_id=str(crm_org_id)

        payload={
            "name": person_name or "Contato Comercial",
            "phone": [{"value": p, "primary": i == 0} for i,p in enumerate(phones)]
        }
        if org_id:
            payload["org_id"]=int(org_id)

        person=pd("POST","persons",json=payload) or {}
        new_id=int(person.get("id") or 0)
        if new_id:
            pd("PUT",f"deals/{deal_id}",json={"person_id":new_id})
            created+=1
            print("OK", deal_id, "person", new_id, "phones", ",".join(phones))
        else:
            errors+=1
            print("ERRO_CREATE", deal_id)

        time.sleep(0.2)

    except Exception as e:
        errors+=1
        print("ERRO", deal_id, e)

print("[FINAL_COUNTS]")
print("created_person_with_phone=", created)
print("skipped_existing_person=", skipped_existing)
print("skipped_no_phone=", skipped_no_phone)
print("errors=", errors)
