from __future__ import annotations

import json
import re
import time
from collections import Counter
from pathlib import Path
from typing import Any, Dict, Iterable, List
from urllib.parse import urlparse

import openpyxl
import requests

from enrich_missing_cnpj_from_xlsx import normalize_cnpj, normalize_phone
from prepare_deals11_writeback import clean_phone, normalize_key


INPUT_XLSX = Path(r"C:\Users\Asus\legacy\bot_sdr_ai\data\deals_pronto_writeback_20260407_200330.xlsx")
OUTPUT_XLSX = Path(r"C:\Users\Asus\legacy\bot_sdr_ai\data\deals_pronto_writeback_phones_web.xlsx")
CACHE_JSON = Path(r"C:\Users\Asus\legacy\bot_sdr_ai\data\phone_web_lookup_cache.json")
TIMEOUT_SEC = 10
DELAY_SEC = 0.4
MAX_PAGES_PER_DEAL = 3


BLOCKED_HOST_TOKENS = (
    "google.", "youtube.", "facebook.", "instagram.", "linkedin.", "tiktok.",
    "reclameaqui", "jusbrasil", "escavador", "econodata", "solutudo",
)


def log(message: str) -> None:
    print(str(message), flush=True)


def load_cache() -> Dict[str, Any]:
    if not CACHE_JSON.exists():
        return {}
    try:
        data = json.loads(CACHE_JSON.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def save_cache(cache: Dict[str, Any]) -> None:
    CACHE_JSON.parent.mkdir(parents=True, exist_ok=True)
    CACHE_JSON.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def read_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    ws = wb["writeback"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row_no, values in enumerate(ws.iter_rows(min_row=2), start=2):
        item = {headers[idx]: cell.value for idx, cell in enumerate(values)}
        item["__row_no"] = row_no
        rows.append(item)
    wb.close()
    return headers, rows


def search_html_google(query: str) -> Dict[str, Any]:
    try:
        response = requests.get(
            "https://www.google.com/search",
            params={"q": query, "hl": "pt-BR", "num": "8"},
            headers={
                "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124 Safari/537.36",
                "accept-language": "pt-BR,pt;q=0.9,en;q=0.8",
            },
            timeout=TIMEOUT_SEC,
        )
        return {"status": response.status_code, "html": response.text[:250000]}
    except Exception as exc:
        return {"status": 0, "html": "", "error": str(exc)[:200]}


def search_html_duckduckgo(query: str) -> Dict[str, Any]:
    try:
        response = requests.get(
            "https://duckduckgo.com/html/",
            params={"q": query},
            headers={"user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) bot-sdr-ai/1.0"},
            timeout=TIMEOUT_SEC,
        )
        return {"status": response.status_code, "html": response.text[:250000]}
    except Exception as exc:
        return {"status": 0, "html": "", "error": str(exc)[:200]}


def extract_urls(html: str) -> List[str]:
    urls: List[str] = []
    seen = set()
    for raw in re.findall(r'href=["\']([^"\']+)["\']', str(html or "")):
        url = raw
        if raw.startswith("/url?") and "q=" in raw:
            match = re.search(r"[?&]q=([^&]+)", raw)
            if match:
                from urllib.parse import unquote
                url = unquote(match.group(1))
        if not url.startswith(("http://", "https://")):
            continue
        host = urlparse(url).netloc.lower()
        if any(token in host for token in BLOCKED_HOST_TOKENS):
            continue
        if url not in seen:
            seen.add(url)
            urls.append(url)
    return urls


def fetch_url(url: str) -> Dict[str, Any]:
    try:
        response = requests.get(
            url,
            headers={
                "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124 Safari/537.36",
                "accept-language": "pt-BR,pt;q=0.9,en;q=0.8",
            },
            timeout=TIMEOUT_SEC,
            allow_redirects=True,
        )
        text = response.text[:300000] if "text" in response.headers.get("content-type", "") else ""
        return {"status": response.status_code, "text": text, "url": response.url}
    except Exception as exc:
        return {"status": 0, "text": "", "url": url, "error": str(exc)[:200]}


def extract_phones(text: str) -> List[str]:
    phones: List[str] = []
    for match in re.findall(r"(?:\+?55\s*)?(?:\(?\d{2}\)?\s*)?\d{4,5}[-.\s]?\d{4}", str(text or "")):
        phone = clean_phone(match)
        if phone and phone not in phones:
            phones.append(phone)
    return phones


def relevant_text(text: str, name: str, cnpj: str) -> bool:
    clean = normalize_key(text[:120000])
    if cnpj and cnpj in re.sub(r"\D+", "", text):
        return True
    name_key = normalize_key(name)
    if name_key and len(name_key) > 8 and name_key in clean:
        return True
    tokens = [token for token in re.findall(r"[a-z0-9]{4,}", normalize_key(name)) if token not in {"ltda", "brasil", "comercio"}]
    hits = sum(1 for token in tokens[:5] if token in clean)
    return bool(tokens and hits >= min(2, len(tokens)))


def choose_phone(candidates: List[str]) -> str:
    if not candidates:
        return ""
    unique = list(dict.fromkeys(clean_phone(item) for item in candidates if clean_phone(item)))
    mobiles = [phone for phone in unique if len(phone) == 11 and phone[2] == "9"]
    return (mobiles or unique or [""])[0]


def lookup_phone(row: Dict[str, Any], cache: Dict[str, Any]) -> Dict[str, Any]:
    deal_id = str(row.get("deal_id") or "")
    if deal_id in cache:
        return dict(cache[deal_id] or {})
    name = str(row.get("nome_final") or "").strip()
    cnpj = normalize_cnpj(row.get("cnpj"))
    queries = [
        f'"{name}" telefone',
        f'"{name}" CNPJ telefone',
    ]
    if cnpj:
        queries.insert(0, f'"{cnpj}" telefone')
    accepted: List[str] = []
    sources: List[str] = []
    statuses: List[str] = []
    for query in queries:
        for engine, search_fn in (("google", search_html_google), ("duckduckgo", search_html_duckduckgo)):
            search = search_fn(query)
            statuses.append(f"{engine}:{search.get('status')}")
            html = str(search.get("html") or "")
            if relevant_text(html, name, cnpj):
                accepted.extend(extract_phones(html))
                sources.append(f"{engine}:snippet")
            urls = extract_urls(html)[:MAX_PAGES_PER_DEAL]
            for url in urls:
                page = fetch_url(url)
                text = str(page.get("text") or "")
                if relevant_text(text, name, cnpj):
                    page_phones = extract_phones(text)
                    if page_phones:
                        accepted.extend(page_phones)
                        sources.append(str(page.get("url") or url)[:200])
            if accepted:
                break
        if accepted:
            break
        time.sleep(DELAY_SEC)
    phone = choose_phone(accepted)
    payload = {"phone": phone, "sources": sources[:3], "statuses": statuses}
    cache[deal_id] = payload
    save_cache(cache)
    return payload


def main() -> int:
    headers, rows = read_rows(INPUT_XLSX)
    target = [row for row in rows if row.get("acao") in {"UPDATE", "RESTAURAR_VENDAS"} and not clean_phone(row.get("telefone"))]
    cache = load_cache()
    counters: Counter[str] = Counter()
    log(f"[PHONE_WEB_INICIO] targets={len(target)} cache={len(cache)}")
    for idx, row in enumerate(target, start=1):
        result = lookup_phone(row, cache)
        phone = clean_phone(result.get("phone"))
        if phone:
            row["telefone"] = phone
            counters["telefone_web"] += 1
            log(f"[TELEFONE_WEB] {idx}/{len(target)} deal={row.get('deal_id')} phone={phone}")
        else:
            counters["sem_telefone"] += 1
        if idx % 25 == 0:
            log(f"[PROGRESSO] {idx}/{len(target)}")
    wb = openpyxl.load_workbook(INPUT_XLSX)
    ws = wb["writeback"]
    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {header: idx + 1 for idx, header in enumerate(header_row)}
    row_by_no = {row["__row_no"]: row for row in rows}
    for row_no, row in row_by_no.items():
        ws.cell(row=row_no, column=col_map["telefone"], value=row.get("telefone") or "")
    summary = wb["resumo"] if "resumo" in wb.sheetnames else wb.create_sheet("resumo")
    summary.append(["telefone_web", counters["telefone_web"]])
    summary.append(["telefone_web_sem_match", counters["sem_telefone"]])
    wb.save(OUTPUT_XLSX)
    log("[PHONE_WEB_RESUMO] " + " ".join(f"{key}={value}" for key, value in sorted(counters.items())))
    log(f"[EXPORTADO] {OUTPUT_XLSX}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
